import os

from openpyxl import Workbook
from openpyxl import load_workbook

FILENAME = 'html'
XLSXTRAILER = 'xlsx'
HTMLTRAILER = 'html'


def write2excel(list):
    if len(list) == 0:
        return
    file = FILENAME + '.' + XLSXTRAILER

    try:
        wb = load_workbook(file)
    except FileNotFoundError as e:
        wb = Workbook()
        wb.save(file)
    finally:
        wb = load_workbook(file)

    ws = wb.active

    for i in list:
        ws.append([i])

    wb.save(file)


def write2html(list):
    if len(list) == 0:
        return
    file = FILENAME + '.' + HTMLTRAILER
    with open(file, 'a+') as file:
        for i in list:
            file.write('<p><a href="' + i + '">' + i + '</a></p>\n')


if __name__ == '__main__':
    list = ['https://www.douban.com/group/topic/90533182', 'https://www.douban.com/group/topic/905331ee',
            'https://www.douban.com/group/topic/9053rr2']
    write2excel(list)
    write2html(list)
